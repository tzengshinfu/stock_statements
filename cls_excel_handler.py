from openpyxl import Workbook
from openpyxl.compat import range
import os
import tempfile
from typing import Union
from typing import List
from openpyxl import load_workbook


class ClsExcelHandler():
    def __init__(self):
        self._books_path = tempfile.gettempdir()

    def save_book(self, book_path: str):
        """
        儲存活頁簿

        Arguments:
            book_path {str} -- 本機路徑
        """
        self._book.save(book_path)

    def write_to_sheet(self, values: Union[List[List[str]], List[str], str]):
        """
        寫入工作表

        Arguments:
            values {Union[List[List[str]], List[str], str]} -- 要寫入的值
        """
        if isinstance(values, list):
            if len(values) > 0:
                if isinstance(values[0], list):
                    for currentIndex in range(0, len(values)):
                        self._sheet.append(values[currentIndex])
                else:
                    self._sheet.append(values)
        elif isinstance(values, str):
            self._sheet.append(values)
        else:
            raise ValueError('values型別只能是(list[list[str]]/list[str]/str)其中之一')

    def open_books_directory(self, books_path: str):
        """
        開啟活頁簿預設儲存目錄

        Arguments:
            books_path {str} -- 本機路徑
        """
        if not os.path.exists(books_path):
            os.makedirs(books_path)
        self._books_path = books_path

    def open_book(self, book_path: str):
        """
        開啟活頁簿(不存在則先建立)

        Arguments:
            book_path {str} -- 本機路徑
        """
        if not self.is_book_existed(book_path):
            self._book = Workbook()
        else:
            self._book = load_workbook(book_path)
        self._sheet = self._book.active

    def open_sheet(self, sheet_name: str):
        """
        開啟工作表(不存在則先建立)

        Arguments:
            sheet_name {str} -- 工作表名稱
        """
        if not self.is_sheet_existed(sheet_name):
            self._book.create_sheet(sheet_name)
        self._sheet = self._book.get_sheet_by_name(sheet_name)
        self._book.active = self._sheet

    def is_book_existed(self, book_path: str) -> bool:
        """
        判斷活頁簿是否存在

        Arguments:
            book_path {str} -- 本機路徑

        Returns:
            bool -- 回傳結果
        """
        return os.path.exists(book_path)

    def is_sheet_existed(self, sheet_name: str) -> bool:
        """
        判斷工作表是否存在

        Arguments:
            sheet_name {str} -- 工作表名稱

        Returns:
            bool -- 回傳結果
        """
        if sheet_name in self._book.sheetnames:
            return True
        else:
            return False
